import openpyxl
from flask import Blueprint, render_template, request, redirect, url_for, flash, session,send_file
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from . import db
import os
from .models import User, Feedback, Project, Step, Role, projectAssigned
from flask_login import login_user, logout_user, login_required
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension
from openpyxl.utils import get_column_letter
from functools import wraps
from .send_email import send_email


def columnadjuster(ws):
    dimholder1 = DimensionHolder(worksheet=ws)
    for row in range(ws.min_row, ws.max_row + 1):
        dimholder1[row] = RowDimension(ws, min=row, max=row, height=40)
        ws.row_dimensions = dimholder1
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=50)
        ws.column_dimensions = dim_holder


def heightadjuster(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True,)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.font = Font(size=22)
basedirauth = os.path.abspath(os.path.dirname(__file__))
auth = Blueprint('auth', __name__)


def access_required(role="ANY"):
    def wrapper(fn):
        @wraps(fn)
        def decorated_view(*args, **kwargs):
            if session.get('role')  not in role or role == "ANY":
                flash('You do not have access to this page')
                return redirect(url_for('auth.login'))
            return fn(*args, **kwargs)
        return decorated_view
    return wrapper

@auth.route('/admin')
@access_required(role="Admin")
def admin():
    return render_template('admin.html', user=User.query.all(),project=Project.query.filter_by(projectApproved=False).all())

@auth.route('/addUser')
@access_required(role="Admin")
def addUser():
    return render_template('addUser.html')

@auth.route('/addUser', methods=['POST'])
@access_required(role="Admin")
def addUser_post():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        user = User.query.filter_by(email=email).first()
        if user:
            flash('Email address already exists')
            return redirect(url_for('auth.addUser'))
        new_user = User(email=email, password=generate_password_hash(password, method='sha256'),type=role,email_confirmed_at=datetime.now())
        new_user.roles.append(Role(name=role))
        db.session.add(new_user)
        db.session.commit()
    return redirect(url_for('auth.addUser'))

@auth.route('/assignProject')
@login_required
@access_required(role="Admin")
def assignProject():
    users = User.query.filter_by(type="ProjectManager").all()
    print(users)
    if users:
        return render_template('assignProject.html', users=users)
    flash('No Project Manager  or no Project Found')
    return render_template('index.html')

@auth.route('/assignProject', methods=['POST'])
@login_required
@access_required(role="Admin")
def assignProject_post():
    if request.method == 'POST':
        projectNo = request.form.get('projectNo')
        projectManager = request.form.get('projectManager')
        project = projectAssigned(projectNo=projectNo, projectManager=projectManager)
        project1 = projectAssigned(projectNo=projectNo, projectManager=session['user'])
        db.session.add(project)
        db.session.add(project1)
        db.session.commit()
    flash('Project Assigned Successfully')
    return redirect(url_for('main.index'))


@auth.route('/approveProject/<projectNo>', methods=['POST'])
@login_required
@access_required(role="Admin")
def approveProject(projectNo):
    project = Project.query.filter_by(projectNo=projectNo).first()
    project.projectApproved = True
    project.projectRejected = False
    db.session.commit()
    stepCount1 = int(project.stepCount)
    partNo = project.partNo
    partName = project.partName
    poQty = project.poQty
    projectStart = project.projectStart
    dispatchDate = project.dispatchDate
    partImage = project.partImage
    customerName = project.customerName
    buyerName = project.buyerName
    projectCoordinator = project.projectCoordinator
    if project.projectApproved == True and project.projectRejected == False:
        createexcel(stepCount1, projectNo, partNo, partName, poQty, projectStart, dispatchDate, partImage, customerName,
                    buyerName, projectCoordinator)
        for i in range(0, stepCount1 + 1):
            step = Step.query.filter_by(projectNo=projectNo, index=i).first()
            if step:
                stepName = step.stepName
                stepDays = step.stepDays
                stepStart = step.stepStart
                stepEnd = step.stepEnd
                stepStatus = step.stepStatus
                create_excel_step(i, stepName, stepDays, stepStart, stepEnd, stepStatus, projectNo)
        # send_email(projectCoordinator, projectNo, partNo, partName, poQty, projectStart, dispatchDate, partImage,customerName, buyerName)
    return redirect(url_for('auth.admin'))

@auth.route('/rejectProject/<projectNo>', methods=['POST'])
@access_required(role="Admin")
@login_required
def rejectProject(projectNo):
    project = Project.query.filter_by(projectNo=projectNo).first()
    db.session.delete(project)
    db.session.commit()
    return redirect(url_for('auth.admin'))


@auth.route('/login')
def login():
    session['role'] = 'ANY'
    session.permanent = True
    return render_template('login.html')

@auth.route('/login', methods=['POST'])
def login_post():
    email = request.form.get('email')
    password = request.form.get('password')
    remember = True if request.form.get('remember') else False
    user = User.query.filter_by(email=email).first()
    # check if the user actually exists
    # take the user-supplied password, hash it, and compare it to the hashed password in the database
    if not user or not check_password_hash(user.password, password):
        flash('Please check your login details and try again.')
        return redirect(url_for('auth.login'))
    role = Role.query.filter_by(id=user.id).first()
    session['role'] = str(role.name)
    session['user'] = user.email
    login_user(user, remember=remember)
    return render_template('index.html')

@auth.route('/signup')
def signup():
    if not User.query.filter(User.email == 'admin@example.com').first():
        superuser = User(email='admin@example.com', password=generate_password_hash('password', method='sha256'),email_confirmed_at=datetime.utcnow(),type='Admin')
        superuser.roles.append(Role(name='Admin'))
        db.session.add(superuser)
        db.session.commit()
    return render_template('signup.html')

@auth.route('/signup', methods=['POST'])
def signup_post():
    # code to validate and add user to database goes here
    email = request.form.get('email')
    role = request.form.get('role')
    password = request.form.get('password')
    user = User.query.filter_by(email=email).first()  # if this returns a user, then the email already exists in database
    if user: # if a user is found, we want to redirect back to signup page so user can try again
        flash('Email address already exists')
        return redirect(url_for('auth.signup'))
    # create a new user with the form data. Hash the password so the plaintext version isn't saved.
    new_user = User(email=email,password=generate_password_hash(password, method='sha256'),email_confirmed=False,email_confirmed_at=datetime.utcnow(),type="Employee")
    new_user.roles.append(Role(name=role))
    # add the new user to the database
    db.session.add(new_user)
    db.session.commit()

    return redirect(url_for('auth.login'))

indexforfeedback = 2
@auth.route('/feedback',methods=['POST','GET'])
@login_required
def feedback():
    if request.method=='POST':
        # feedbackId = request.form['feedbackId']
        projectNo = request.form['projectNo']
        email = request.form['email']
        password = request.form['password']
        curDate = request.form['curDate']
        curDate = datetime.strptime(curDate, '%Y-%m-%d')
        remarks = request.form['remarks']
        role=request.form['role']
        feedbackImage = request.files['feedbackImage']
        project = Project.query.filter_by(projectNo=projectNo).first()
        if project is None:
            return render_template('feedback.html', message='please enter correct project number')
        if email==''or password=='' or projectNo=='':
            return render_template('feedback.html', message='please enter required field')
        user=User.query.filter_by(email=email).first()
        checkpw=check_password_hash(user.password,password)
        if user is None or checkpw!=True or role!=user.type:
            return render_template('feedback.html', message='please enter correct email or password')
        if user.projectAssigned!=projectNo:
            return render_template('feedback.html', message='you are not assigned to this project')
        feedback = Feedback.query.filter_by(projectNo=projectNo).first()
        maxfeedbackId = db.session.query(db.func.max(Feedback.indexforFeedback)).scalar()
        if feedback is None:
            feedbackId = projectNo + '-1'
            indexforfeedback = 2
        else:
            feedbackId = projectNo + '-' + str(maxfeedbackId)
            indexforfeedback = maxfeedbackId + 1
        if feedbackImage:
            ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg', 'gif'])
            def allowed_file(filename):
                return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
            def saveimage():
                if feedbackImage and allowed_file(feedbackImage.filename):
                    filenameoffeedbackimage = secure_filename(feedbackImage.filename)
                    feedbackImage.save(os.path.join(basedirauth, 'images' , filenameoffeedbackimage))
            saveimage()
            data = Feedback(feedbackId=feedbackId, projectNo=projectNo, email=email,password=generate_password_hash(password, method='sha256'),role=role, curDate=curDate,remarks=remarks, feedbackImage=feedbackImage.read(),indexforFeedback=indexforfeedback,filename=feedbackImage.filename)
        else:
            data = Feedback(feedbackId=feedbackId, projectNo=projectNo, email=email,password=generate_password_hash(password, method='sha256'), role=role,curDate=curDate ,remarks=remarks,indexforFeedback=indexforfeedback)
        db.session.add(data)
        db.session.commit()
        def createExcel(indexforfeedback):
            wb = load_workbook(os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx'))
            ws = wb.active
            ws.cell(row=15,column=indexforfeedback).value = curDate
            ws.merge_cells(str(get_column_letter(indexforfeedback))+'16'+':'+str(get_column_letter(indexforfeedback))+'21')
            ws.cell(row=16,column=indexforfeedback).value = remarks
            ws.merge_cells(str(get_column_letter(indexforfeedback))+'22'+':'+str(get_column_letter(indexforfeedback))+'27')
            if feedbackImage:
                img = openpyxl.drawing.image.Image(
                    os.path.join(basedirauth, 'images' , secure_filename(feedbackImage.filename)))
                img.height = 280
                img.width = 460
                img.anchor = str(get_column_letter(indexforfeedback))+'22'
                ws.add_image(img)
            columnadjuster(ws)
            heightadjuster(ws)
            for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(vertical='top')
            wb.save(os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx'))
        createExcel(indexforfeedback)
        # send_email(email, projectNo, curDate, role, remarks, basedirauth)
        return render_template('feedback.html', message='Feedback submitted successfully')

@auth.route('/createProject',methods=['POST'])
@login_required
@access_required('Admin or ProjectManager')
def createProject():
    if request.method=='POST':
        projectNo=request.form['projectNo']
        customerName=request.form['customerName']
        buyerName=request.form['buyerName']
        projectCoordinator=request.form['projectCoordinator']
        projectStart=request.form['projectStart']
        dispatchDate=request.form['dispatchDate']
        poQty=request.form['poQty']
        partNo=request.form['partNo']
        partName=request.form['partName']
        partImage=request.files['partImage']
        projectStart = datetime.strptime(projectStart, '%Y-%m-%d')
        dispatchDate = datetime.strptime(dispatchDate, '%Y-%m-%d')
        stepCount = request.form['stepCount']
        stepCount1 = int(stepCount)
        employees = request.form.getlist('employee[]')
        if projectNo =='':
            flash('No project has been assigned to the current session user')
            return redirect(url_for('main.create_project'))
        if  customerName=='' or buyerName=='' or projectCoordinator=='' or dispatchDate=='' or poQty=='' or partNo=='' or partName=='' or projectStart=='' or stepCount=='' or employees=='':
            flash('Please enter required fields')
            return redirect(url_for('main.create_project'))
        project = Project.query.filter_by(projectNo=projectNo).first()
        if project:
            flash('Project already exists')
            return redirect(url_for('main.create_project'))
        if partImage:
            ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg', 'gif'])
            def allowed_file(filename):
                return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
            def saveimage():
                if partImage and allowed_file(partImage.filename):
                    filenameofpartimage = secure_filename(partImage.filename)
                    partImage.save(os.path.join(basedirauth, 'images' , filenameofpartimage))
            saveimage()
            data1 = Project(projectNo=projectNo, customerName=customerName, buyerName=buyerName,
                            projectCoordinator=projectCoordinator, projectStart=projectStart, dispatchDate=dispatchDate,
                            poQty=poQty, stepCount=stepCount1, partNo=partNo, partName=partName,
                            partImage=partImage.read(),usersAssigned=str(employees))
        else:
            data1 = Project(projectNo=projectNo, customerName=customerName, buyerName=buyerName,
                            projectCoordinator=projectCoordinator, projectStart=projectStart, dispatchDate=dispatchDate,
                            poQty=poQty, stepCount=stepCount1, partNo=partNo, partName=partName,usersAssigned=str(employees))
        for employee in employees:
            employee.split(',')
            user = User.query.filter_by(id=employee).first()
            user.projectAssigned = projectNo
        #access to the database
        # def createexcel(stepCount1,projectNo,partNo,partName,poQty,projectStart,dispatchDate,partImage,customerName,buyerName,projectCoordinator):
        #     wb = Workbook()
        #     ws = wb.active
        #     ws.title = "Project Status"
        #     if int(stepCount)>4:
        #         ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2+stepCount1)
        #     else:
        #         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        #     wscsr=ws.cell(row=1,column=1)
        #     wscsr.value = 'Project Plan/Status'
        #     wscsr.alignment = Alignment(horizontal='center', vertical='center')
        #     ws['A2']="Customer Name"
        #     ws['A3']="Part/Item Name"
        #     ws['A4']="Part/Item No."
        #     ws['A5']="Buyer Name"
        #     ws['A6']="ACT Project Coordinator"
        #     ws['B2']=customerName
        #     ws['B3']=partName
        #     ws['B4']=partNo
        #     ws['B5']=buyerName
        #     ws['B6']=projectCoordinator
        #     ws['C2']="Project No."
        #     ws['C3']="Project Start Date"
        #     ws['C4']="Dispatch Date"
        #     ws['C5']="PO Qty"
        #     ws['C6']="status as on"
        #     ws['D2']=projectNo
        #     ws['D3']=projectStart
        #     ws['D4']=dispatchDate
        #     ws['D5']=poQty
        #     ws['D6']=datetime.now()
        #     ws.merge_cells('E2:E6')
        #     # im1= Image.save(partImage,os.path.join(basedirauth, 'static/images/' + partNo + '.png'))
        #     if partImage:
        #         img = openpyxl.drawing.image.Image(os.path.join(basedirauth, 'images/' + secure_filename(partImage.filename)))
        #         img.height = 280
        #         img.width = 460
        #         img.anchor = 'E2'
        #         ws.add_image(img)
        #     ws['A8']="Scheduled Qty"
        #     ws.merge_cells(start_row=9, start_column=1, end_row=13, end_column=1)
        #     wscsr1 = ws.cell(row=9, column=1)
        #     wscsr1.value = poQty
        #     wscsr1.alignment = Alignment(horizontal='center', vertical='center')
        #     ws['B9']="Days"
        #     ws['B10']="Start"
        #     ws['B11']="End"
        #     ws['B12']="Status"
        #     ws['B13']="Actual completion date"
        #     ws['A15']="Status on"
        #     ws['A16']="Remarks"
        #     ws['A22']="Image"
        #     excelname= os.path.join(basedirauth, 'worksheets/'+projectNo+'.xlsx')
        #     heightadjuster(ws)
        #     columnadjuster(ws)
        #     wb.save(excelname)
        db.session.add(data1)
        db.session.commit()
        return render_template('steps.html',projectNo=projectNo,stepCount=stepCount1)


def createexcel(stepCount1, projectNo, partNo, partName, poQty, projectStart, dispatchDate, partImage, customerName,
                buyerName, projectCoordinator):
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Status"
    if stepCount1 > 4:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + stepCount1)
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wscsr = ws.cell(row=1, column=1)
    wscsr.value = 'Project Plan/Status'
    wscsr.alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'] = "Customer Name"
    ws['A3'] = "Part/Item Name"
    ws['A4'] = "Part/Item No."
    ws['A5'] = "Buyer Name"
    ws['A6'] = "ACT Project Coordinator"
    ws['B2'] = customerName
    ws['B3'] = partName
    ws['B4'] = partNo
    ws['B5'] = buyerName
    ws['B6'] = projectCoordinator
    ws['C2'] = "Project No."
    ws['C3'] = "Project Start Date"
    ws['C4'] = "Dispatch Date"
    ws['C5'] = "PO Qty"
    ws['C6'] = "status as on"
    ws['D2'] = projectNo
    ws['D3'] = projectStart
    ws['D4'] = dispatchDate
    ws['D5'] = poQty
    ws['D6'] = datetime.now()
    ws.merge_cells('E2:E6')
    # im1= Image.save(partImage,os.path.join(basedirauth, 'static/images/' + partNo + '.png'))
    if partImage:
        img = openpyxl.drawing.image.Image(os.path.join(basedirauth, 'images' , secure_filename(partImage.filename)))
        img.height = 280
        img.width = 460
        img.anchor = 'E2'
        ws.add_image(img)
    ws['A8'] = "Scheduled Qty"
    ws.merge_cells(start_row=9, start_column=1, end_row=13, end_column=1)
    wscsr1 = ws.cell(row=9, column=1)
    wscsr1.value = poQty
    wscsr1.alignment = Alignment(horizontal='center', vertical='center')
    ws['B9'] = "Days"
    ws['B10'] = "Start"
    ws['B11'] = "End"
    ws['B12'] = "Status"
    ws['B13'] = "Actual completion date"
    ws['A15'] = "Status on"
    ws['A16'] = "Remarks"
    ws['A22'] = "Image"
    excelname = os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx')
    heightadjuster(ws)
    columnadjuster(ws)
    wb.save(excelname)


def create_excel_step(i, stepName, stepDays, stepStart, stepEnd, stepStatus,projectNo):
    wb = load_workbook(os.path.join(basedirauth, 'worksheets', projectNo + '.xlsx'))
    ws = wb.active
    ws.cell(column=2 + i, row=8).value = stepName
    ws.cell(column=2 + i, row=9).value = stepDays
    ws.cell(column=2 + i, row=10).value = stepStart
    ws.cell(column=2 + i, row=11).value = stepEnd
    ws.cell(column=2 + i, row=12).value = stepStatus
    excelname = os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx')
    columnadjuster(ws)
    heightadjuster(ws)
    wb.save(excelname)



@auth.route('/steps/<projectNo>/<int:stepCount>',methods=['POST','GET'])
@login_required
@access_required('Admin or ProjectManager')
def steps(projectNo,stepCount):
    if request.method=='POST':
        for i in range(1,stepCount+1):
            stepName=request.form['stepName'+str(i)]
            stepDays=request.form['stepDays'+str(i)]
            stepStart=request.form['stepStart'+str(i)]
            stepEnd=request.form['stepEnd'+str(i)]
            stepStatus=request.form['stepStatus'+str(i)]
            stepStart = datetime.strptime(stepStart, '%Y-%m-%d')
            stepEnd = datetime.strptime(stepEnd, '%Y-%m-%d')
            data2=Step(projectNo=projectNo,stepName=stepName,stepDays=stepDays,stepStart=stepStart,stepEnd=stepEnd,stepStatus=stepStatus,stepActualCompletion=stepEnd,index=i)
            db.session.add(data2)
            db.session.commit()
            # def create_excel_step(i,stepName, stepDays, stepStart, stepEnd, stepStatus):
            #     wb = load_workbook(os.path.join(basedirauth, 'worksheets/'+projectNo+'.xlsx'))
            #     ws = wb.active
            #     ws.cell(column=2+i,row=8).value=stepName
            #     ws.cell(column=2+i,row=9).value=stepDays
            #     ws.cell(column=2+i,row=10).value=stepStart
            #     ws.cell(column=2+i,row=11).value=stepEnd
            #     ws.cell(column=2+i,row=12).value=stepStatus
            #     excelname = os.path.join(basedirauth, 'worksheets/' + projectNo + '.xlsx')
            #     columnadjuster(ws)
            #     heightadjuster(ws)
            #     wb.save(excelname)
            project=Project.query.filter_by(projectNo=projectNo).first()
            if project.projectApproved == True and project.projectRejected == False:
                create_excel_step(i,stepName, stepDays, stepStart, stepEnd, stepStatus)
        return redirect(url_for('main.index'))

@auth.route('/project_details')
@login_required
@access_required('Admin or ProjectManager')
def project_details():
    return render_template('project_details.html',projectList=Project.query.all())

@auth.route('/project_details/<projectNo>')
@login_required
@access_required('Admin or ProjectManager')
def project_details1(projectNo):
    return render_template('project_details1.html',basedirauth=basedirauth,feedback=Feedback.query.filter_by(projectNo=projectNo).all())

@auth.route('/downloadWorksheet/<projectNo>', methods=['POST', 'GET'])
@login_required
@access_required('Admin or ProjectManager')
def downloadWorksheet(projectNo):
    file_path = os.path.join(basedirauth, 'worksheets', projectNo + '.xlsx')

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        flash('File not found')
        return render_template('project_details.html',projectList=Project.query.all())

@auth.route('/updatefeedback/<projectNo>',methods=['POST','GET'])
@login_required
@access_required('Admin or ProjectManager')
def updatefeedback(projectNo):
    if request.method=='POST':
        project = Project.query.filter_by(projectNo=projectNo).first()
        if project is None:
            flash('No such project exists')
            return redirect(url_for('main.index'))
        feedbackId=request.form['feedbackId']
        curDate = request.form['curDate']
        curDate = datetime.strptime(curDate, '%Y-%m-%d')
        remarks=request.form['remarks']
        feedbackImage=request.files['feedbackImage']
        feedback= Feedback.query.filter_by(feedbackId=feedbackId).first()
        feedback.curDate=curDate
        feedback.remarks=remarks
        if feedbackImage:
            ALLOWED_EXTENSIONS = set(['png', 'jpg', 'jpeg', 'gif'])
            def allowed_file(filename):
                return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
            def saveimage():
                if feedbackImage and allowed_file(feedbackImage.filename):
                    filenameoffeedbackimage = secure_filename(feedbackImage.filename)
                    feedbackImage.save(os.path.join(basedirauth, 'images' , filenameoffeedbackimage))
            saveimage()
        feedback.feedbackImage=feedbackImage.read()
        db.session.commit()
        indexforfeedback=Feedback.indexforfeedback
        def updateexcel(indexforfeedback):
            wb = load_workbook(os.path.join(basedirauth, 'worksheets',projectNo+'.xlsx'))
            ws = wb.active
            ws.cell(row=15, column=indexforfeedback).value = curDate
            ws.merge_cells(
                str(get_column_letter(indexforfeedback)) + '16' + ':' + str(get_column_letter(indexforfeedback)) + '21')
            ws.cell(row=16, column=indexforfeedback).value = remarks
            ws.merge_cells(
                str(get_column_letter(indexforfeedback)) + '22' + ':' + str(get_column_letter(indexforfeedback)) + '27')
            if feedbackImage:
                img = openpyxl.drawing.image.Image(
                    os.path.join(basedirauth, 'images' , secure_filename(feedbackImage.filename)))
                img.height = 280
                img.width = 460
                img.anchor = str(get_column_letter(indexforfeedback)) + '22'
                ws.add_image(img)
            columnadjuster(ws)
            heightadjuster(ws)
            for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(vertical='top')
            wb.save(os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx'))
        updateexcel(indexforfeedback)
        return redirect(url_for('main.index'))

@auth.route('/editproject/<projectNo>',methods=['POST','GET'])
@login_required
@access_required('Admin or ProjectManager')
def editproject(projectNo):
    if request.method=='GET':
        return render_template('editproject.html',project=Project.query.filter_by(projectNo=projectNo).first())
    if request.method=='POST':
        CustomerName=request.form['CustomerName']
        buyerName=request.form['buyerName']
        partName=request.form['partName']
        partNo=request.form['partNo']
        projectCoordinator=request.form['projectCoordinator']
        projectStart=request.form['projectStart']
        dispatchDate=request.form['dispatchDate']
        poQty=request.form['poQty']
        stepCount=request.form['stepCount']
        partImage=request.files['partImage']
        projectStart = datetime.strptime(projectStart, '%Y-%m-%d')
        dispatchDate = datetime.strptime(dispatchDate, '%Y-%m-%d')
        data = Project.query.filter_by(projectNo=projectNo).first()
        def update(column):
            if column:
                data.column=column
        update(CustomerName)
        update(buyerName)
        update(partName)
        update(partNo)
        update(projectCoordinator)
        update(projectStart)
        update(dispatchDate)
        update(poQty)
        update(stepCount)
        db.session.commit()
        stepCount1=int(stepCount)
        def update_excel():
            wb = load_workbook(os.path.join(basedirauth, 'worksheets',projectNo+'.xlsx'))
            ws = wb.active
            ws.cell(row=2, column=2).value = CustomerName
            ws.cell(row=3, column=2).value = buyerName
            ws.cell(row=4, column=2).value = partName
            ws.cell(row=5, column=2).value = partNo
            ws.cell(row=6, column=2).value = projectCoordinator
            ws.cell(row=2, column=4).value = projectNo
            ws.cell(row=3, column=4).value = projectStart
            ws.cell(row=4, column=4).value = dispatchDate
            ws.cell(row=5, column=4).value = poQty
            ws.cell(row=6, column=4).value = stepCount1
            columnadjuster(ws)
            heightadjuster(ws)
            wb.save(os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx'))
        update_excel()
        return render_template('steps.html',projectNo=projectNo,stepCount=stepCount1)

@auth.route('/stepStatus/<projectNo>')
@login_required
@access_required('Admin or ProjectManager')
def stepStatus(projectNo):
    return render_template('stepStatus.html',steps=Step.query.filter_by(projectNo=projectNo).all(),projectNo=projectNo)

@auth.route('/stepStatus_post/<projectNo>',methods=['POST'])
@login_required
@access_required('Admin or ProjectManager')
def stepStatus_post(projectNo):
    if request.method=='POST':
        stepStatus=request.form['stepStatus']
        stepIndex=request.form['stepIndex']
        stepIndex=int(stepIndex)
        step = Step.query.filter_by(projectNo=projectNo, index=stepIndex).first()
        if stepStatus=='' or stepIndex=='':
            flash('Please fill all the fields')
            return redirect(url_for('auth.stepStatus',steps=Step.query.filter_by(projectNo=projectNo).all(),projectNo=projectNo))
        if stepStatus=='COMP':
            stepCompletionDate=datetime.now()
            print(stepCompletionDate)
            step.stepActualCompletion = stepCompletionDate
        else:
            step.stepActualCompletion = None
            stepCompletionDate = 'x-x-x'
        step.stepStatus=stepStatus
        db.session.commit()
        def updateexcel(stepCompletionDate):
            wb = load_workbook(os.path.join(basedirauth, 'worksheets',projectNo+'.xlsx'))
            ws = wb.active
            ws.cell(row=12, column=2+stepIndex).value = stepStatus
            ws.cell(row=13, column=2+stepIndex).value = stepCompletionDate
            columnadjuster(ws)
            heightadjuster(ws)
            wb.save(os.path.join(basedirauth, 'worksheets' , projectNo + '.xlsx'))
        updateexcel(stepCompletionDate)
        return render_template('stepStatus.html',steps=Step.query.filter_by(projectNo=projectNo).all(),projectNo=projectNo)

@auth.route('/deleteProject/<projectNo>',methods=['POST','GET'])
@login_required
@access_required('Admin or ProjectManager')
def deleteProject(projectNo):
    alert = "Are you sure you want to delete this project?"
    if request.method=='POST':
        project = Project.query.filter_by(projectNo=projectNo).first()
        feedback = Feedback.query.filter_by(projectNo=projectNo).all()
        if feedback:
            db.session.delete(feedback)
        db.session.delete(project)
        db.session.commit()
        return redirect(url_for('main.index'))
    return redirect(url_for('main.index'))

@auth.route('/deleteFeedback/<feedbackId>',methods=['POST','GET'])
@login_required
@access_required('Admin' or 'Project Manager')
def deleteFeedback(feedbackId):
    alert = "Are you sure you want to delete this feedback?"
    if request.method=='POST':
        feedback = Feedback.query.filter_by(feedbackId=feedbackId).first()
        db.session.delete(feedback)
        db.session.commit()
        return redirect(url_for('main.index'))
    return redirect(url_for('main.index'))

@auth.route('/deleteUser/<userId>',methods=['POST','GET'])
@login_required
@access_required('Admin')
def deleteUser(userId):
    alert = "Are you sure you want to delete this user?"
    if request.method=='POST':
        user = User.query.filter_by(id=userId).first()
        db.session.delete(user)
        db.session.commit()
        return redirect(url_for('auth.admin'))
    return redirect(url_for('auth.admin'))

@auth.route('/logout')
@login_required
def logout():
    session.clear()
    logout_user()
    return redirect(url_for('main.index'))
